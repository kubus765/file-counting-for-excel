import os
import openpyxl
import time
import sys
from tkinter import Tk, Checkbutton, Text
from tkinter.filedialog import askopenfilename
import threading
import datetime

def on_window_close():
    print_to_console("Closing")
    time.sleep(0.1)
    window.destroy()
    os._exit(1)
    
def update_excel(file_path):
    # Extract the relevant information from the file name
    file_name = os.path.basename(file_path)
    content = file_name.split('.')[0]  # Assuming the content is before the file extension

    # Define the conditions and their respective positions
    conditions = {
        'ISA00045-01': (1, 2),
        'ISA00045-02': (2, 2),
        'ISA00045-03': (3, 2),
        # Add more conditions as needed
    }
    
    # Check if the serial number has been processed already
    serial_number = get_serial_number(file_name)
    if serial_number in processed_serial_numbers:
        print_to_console(f"Serial number {serial_number} has already been processed. Skipping file.")
        return

    # Attempt to open the Excel spreadsheet with retry logic
    max_retries = 9999
    retries = 0
    while retries < max_retries:
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            break
        except Exception as e:
            print_to_console(f"Error opening the Excel file: {e}")
            print_to_console(" ")
            print_to_console("Retrying, sleep 10 seconds...")
            print_to_console("Please save and close the document!")
            print_to_console(" ")
            retries += 1
            time.sleep(10)  # Delay before retrying

    # Check if the maximum number of retries has been reached
    if retries == max_retries:
        print_to_console("Unable to open the Excel file. Exiting...")
        return
    
    worksheet = workbook.active

    # Check if the condition exists in the file name and the file content contains "PASS"
    for condition, position in conditions.items():
        if condition in content and 'PASS' in open(file_path).read():
            row, column = position
            current_value = worksheet.cell(row=row, column=column).value
            new_value = current_value + 1 if current_value else 1
            worksheet.cell(row=row, column=column, value=new_value)
            break  # Exit the loop if condition is found

    # Attempt to save the changes to the spreadsheet with retry logic
    retries = 0
    while retries < max_retries:
        try:
            workbook.save(excel_file_path)
            break
        except Exception as e:
            print_to_console(f"Error saving the Excel file: {e}")
            print_to_console(" ")
            print_to_console("Retrying, sleep 10 seconds...")
            print_to_console("Please save and close the document!")
            print_to_console(" ")
            retries += 1
            time.sleep(10)  # Delay before retrying

    # Check if the maximum number of retries has been reached
    if retries == max_retries:
        print_to_console("Unable to save the Excel file. Exiting...")
        return

    # Add the serial number to the processed serial numbers set
    processed_serial_numbers.add(serial_number)
    
running = True
# Recursive function to scan all folders within the directory for new text files
def scan_directory():
    while True:
        if not service_mode_enabled:
            for root, dirs, files in os.walk(directory_path):
                for file in files:
                    if file.endswith('.txt'):
                        file_path = os.path.join(root, file)
                        modified_time = os.path.getmtime(file_path)
                        if modified_time > start_time and file_path not in processed_files:
                            update_excel(file_path)
                            processed_files.add(file_path)
                            print_to_console(f"Processed {file_path}.")


        # Add a delay before checking for new files again
        time.sleep(10)  # Delay of 10 seconds

# Extract the serial number from the file name
def get_serial_number(file_name):
    # Modify this function based on the format of the serial number in the file name
    serial_number = file_name.split('_')[1]
    return serial_number



# Service mode checkbox toggle
def toggle_service_mode():
    global service_mode_enabled, start_time

    if service_mode_enabled:
        service_mode_enabled = False
        start_time = time.time()  # Refresh the start time when Service Mode is disabled
        print_to_console_unpause_message()
    else:
        service_mode_enabled = True
        print_to_console_pause_message()

def print_to_console_pause_message():
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print_to_console(f"Script paused at {current_time}.")

def print_to_console_unpause_message():
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print_to_console(f"Script unpaused at {current_time}. Timestamp refreshed.")

def print_to_console(message):
    console_text.config(state='normal')  # Enable editing the console
    console_text.insert('end', message + '\n')  # Append the message to the console
    console_text.config(state='disabled')  # Disable editing the console


try:
    # Display a message prompting the user to select the current spreadsheet
    print("Please choose the current spreadsheet.")

    # Prompt user to select the Excel file, exit if no file is selected
    root = Tk()
    root.withdraw()
    root.title("Please choose the current spreadsheet")
    excel_file_path = askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not excel_file_path:
        print("No Excel file selected. Exiting...")
        sys.exit()

    # Monitor a directory for new text files
    directory_path = 'text_files'
    start_time = time.time()
    processed_files = set()
    processed_serial_numbers = set()
    service_mode_enabled = False

    # Display a message after loading the file
    print("File loaded:", excel_file_path)
    print(" ")

    # Create the main window
    window = Tk()
    window.title("Test Counter App")
    window.geometry("600x300")
    window.resizable(False, False)
    window.protocol("WM_DELETE_WINDOW", on_window_close)
    # Console Text widget
    console_text = Text(window, state='disabled', height=15)
    console_text.pack(fill='both', expand=True, padx=10, pady=10)

    # Service Mode checkbox
    service_mode_checkbox = Checkbutton(window, text="Service Mode", command=toggle_service_mode)
    service_mode_checkbox.pack(side='top', anchor='w')

    # Start a thread to continuously scan the directory for new files
    scan_thread = threading.Thread(target=scan_directory)
    scan_thread.start()

    # Start the Tkinter event loop
    window.mainloop()

except Exception as e:
    print(f"An error occurred: {e}")

